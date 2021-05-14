import unittest
import time
import openpyxl
import os.path
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from data import data as user
from data import locator as lt


class PythonOrgSearch(unittest.TestCase):

    def setUp(self):
        from selenium import webdriver
        self.driver = webdriver.Chrome(ChromeDriverManager().install())

    def test_1_search_in_daraz(self):
        
        dirname = Path(__file__).resolve().parent
        print(dirname)
        file_name = "user.txt"
        file_location = os.path.join(dirname,'data', file_name)
        fd=open(file_location,"r")
        lines=fd.readlines()
        username=lines[0]
        password=lines[1]
        singleproduct=lines[2]
        fd.close()
       

        book = openpyxl.load_workbook('D:\\TestAutomation\\test1\\data\\products.xlsx')
        sheet = book.active
        m_row = sheet.max_row
        plist=[]
        for i in range(2, m_row + 1):
            cell = sheet.cell(row = i, column = 1)
            plist.append(cell.value)

        driver = self.driver
        driver.maximize_window()
        # driver.implicitly_wait(40)
        # driver.set_page_load_timeout(10)
        driver.get('https://www.daraz.com.bd/')
        # driver.get("https://member.daraz.com.bd/user/login?spm=a2a0e.home.header.d5.73524591fPQMqk&redirect=https%3A%2F%2Fwww.daraz.com.bd%2F")

        driver.find_element_by_xpath(lt.loginlink).click()
        time.sleep(5)
        
        driver.find_element_by_xpath(lt.usernameinput).send_keys(user.username)
        driver.find_element_by_xpath(lt.userpasswordinput).send_keys(user.password)
        driver.find_element_by_xpath(lt.loginbtn).click()
        time.sleep(5)
        # print(plist)
        
        for item in plist:     
            qs=driver.find_element_by_name(lt.searchproduct)  
            # qs.clear()
            qs.send_keys(Keys.CONTROL + "a");
            qs.send_keys(Keys.DELETE);
            qs.send_keys(item)
            qs.send_keys(Keys.RETURN)
            
            # time.sleep(3)

            # driver.find_element_by_xpath(lt.brandcheckbox).click()
            # driver.find_element_by_xpath(lt.servicechebox).click()
            time.sleep(5)

            products = driver.find_elements_by_class_name(lt.productlist)

            productsname = []
            for product in products:
                # a=product.text
                productsname.append(product.text)
            # print(len(productsname))
            productsname.append("Number of products: "+ str(len(productsname)))
            # print(productsname)
            dirname = Path(__file__).resolve().parent
            file_name = "Item_" + str(item)+".txt"
            file_location = os.path.join(dirname,'data', file_name)
            # print(filename)
            
            with open(file_location, 'w', encoding="utf-8") as f:
                for row in productsname:
                    f.write(row + '\n')


    def test_2_cartproduct_in_daraz(self):

        
        dirname = Path(__file__).resolve().parent
        print(dirname)
        file_name = "user.txt"
        file_location = os.path.join(dirname,'data', file_name)
        f=open(file_location,"r")
        lines=f.readlines()
        username=lines[0]
        password=lines[1]
        singleproduct=lines[2]
        f.close()
        
        driver =self.driver
        driver.maximize_window()
        # driver.implicitly_wait(40)
        # driver.get('https://www.daraz.com.bd/')
        driver.get("https://member.daraz.com.bd/user/login?spm=a2a0e.home.header.d5.73524591fPQMqk&redirect=https%3A%2F%2Fwww.daraz.com.bd%2F" )

        driver.find_element_by_xpath(lt.usernameinput).send_keys(user.username)
        driver.find_element_by_xpath(lt.userpasswordinput).send_keys(user.password)
        driver.find_element_by_xpath(lt.loginbtn).click()
        time.sleep(5)

        driver.find_element_by_name(lt.searchproduct).send_keys(user.product, Keys.RETURN)
        time.sleep(5)

        driver.find_element_by_xpath(lt.productlink).click()
        time.sleep(3)

        pn = driver.find_element_by_xpath(lt.productname).text
        pb = driver.find_element_by_xpath(lt.productbrand).text
        print(pn,pb, sep='\n')

        driver.find_element_by_xpath(lt.addtocartbtn).click()
        time.sleep(3)
        driver.find_element_by_xpath(lt.gotocartbtn).click()
        time.sleep(5)
        # cpn = driver.find_element_by_class_name(lt.cartproductname).text
        # cpb = driver.find_element_by_class_name(lt.cartproductbrand).text
        cpn = driver.find_element_by_xpath(lt.cartproductname).text
        cpb = driver.find_element_by_xpath(lt.cartproductbrand).text

        print(cpn,cpb, sep='\n')
        self.assertEqual(pn, cpn, "cart product name dont match")
        self.assertIn(pb, cpb, "cart product brand dont match")

    def test_3_affiliate_in_daraz(self):

        driver =self.driver
        driver.maximize_window()
        driver.get('https://www.daraz.com.bd/')

        driver.find_element_by_xpath(lt.affilitatelink).click()
        time.sleep(3)
        x = driver.find_element_by_xpath(lt.affiliatepageurl).text
        # print(x)
        self.assertEqual('www.daraz.com.bd', x, "url is not matching")

        driver.find_element_by_xpath(lt.helpcenterlink).click()
        time.sleep(3)
        self.assertIn("Help Center",driver.title, "link is not valid")

        driver.back()
        time.sleep(3)
        driver.find_element_by_xpath(lt.howtobuylink).click()
        time.sleep(3)
        self.assertIn("Help Center",driver.title, "link is not valid")

        driver.back()
        time.sleep(3)
        driver.find_element_by_xpath(lt.returnndgoodslink).click()
        time.sleep(3)
        self.assertIn("Return Goods",driver.title, "link is not valid")

        driver.back()
        time.sleep(3)
        driver.find_element_by_xpath(lt.contactuslink).click()
        time.sleep(3)
        self.assertIn("Contact Us",driver.title, "link is not valid")

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    unittest.main()
